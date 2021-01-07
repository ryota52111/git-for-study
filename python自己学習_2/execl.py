#ステップ1｜ライブラリ
from openpyxl import load_workbook
import win32com.client
import datetime
import os
from O365 import Account

#ステップ2｜所定フォルダ内の「Book1.xlsm」を指定して読み込む
filepath = r'C:\Users\r_sas\Desktop\Outlook_Book1.xlsx'
wb = load_workbook(filename=filepath)
ws1 = wb['データ']

#ステップ3｜集計範囲の取得
startdate=ws1['B2'].value
enddate=ws1['B3'].value

#ステップ4｜Outlookの情報を取得
credentials = ('7c4579dd-2ad0-426b-a306-e1bb5bb0cc79>', 'm~5XT2Nl9OauYx65V2-UuA0-7QLY2-p4s-')
account = Account(credentials)
account.authenticate(scopes=['basic', 'message_all'])
mailbox = account.mailbox()
inbox = mailbox.inbox_folder()
print(inbox)
outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
inbox = outlook.Folders("r_sasaki_111@outlook.jp").Folders("受信トレイ")
print(inbox)
messages=inbox.Items
print(messages)

#ステップ5｜Outlookのt添付ファイル保管用フォルダの作成
now=datetime.datetime.now()
foldermake=now.strftime("%Y/%m/%d %H:%M:%S")
foldermake = foldermake.replace("/","-")
foldermake = foldermake.replace(":","-")
foldermake = foldermake.replace(" ","-")
newfolder_path = r'C:\Users\r_sas\Desktop\Outlook_'+foldermake
os.makedirs(newfolder_path)

#ステップ6｜Outlookの情報を取得
a=0
for message in messages:
    RT=message.ReceivedTime
    hiduketime = datetime.datetime(RT.year ,RT.month, RT.day, RT.hour, RT.minute, RT.second)
    if str(startdate) <= str(hiduketime) <= str(enddate):  
        ws1.cell(row=7+a, column=1).value = a+1
        ws1.cell(row=7+a, column=2).value = str(message.Subject)
        ws1.cell(row=7+a, column=3).value = hiduketime
        ws1.cell(row=7+a, column=4).value = str(message.Sender)
        ws1.cell(row=7+a, column=5).value = str(message.body[0:100])
        if message.Attachments.Count > 0:
            myDate = RT.strftime("%Y/%m/%d %H:%M:%S")
            myDate = myDate.replace("/","-")
            myDate = myDate.replace(":","-")
            myDate = myDate.replace(" ","-")
            datefolder_path = newfolder_path + '\\' + myDate
            os.makedirs(datefolder_path)
            for myAttachment in message.Attachments:
                print(myAttachment.FileName)
                myAttachment.SaveAsFile(datefolder_path + '\\' + myAttachment.FileName)
            ws1.cell(row=7+a, column=6).value = '有'
        else:
            ws1.cell(row=7+a, column=6).value = '無'
        a=a+1

#ステップ7｜「Outlook_Book2.xlsx」として所定のフォルダに保存する
newfilepath = 'C:/Users/---/Outlook_Book2.xlsx'
wb.save(newfilepath)
print('保存しました')